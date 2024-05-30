from openpyxl import load_workbook

##### CONF VAR ######

list_to_search = ["find_me_if_you_can", "yolo"]

#####################

wb = load_workbook('example.xlsx')

l_sheet = wb.sheetnames

l_all = []

l_limit = []

l_limitc = []

l_limitr = []

for i in range(0, len(l_sheet)):

    ws = wb[l_sheet[i]]

    min_row = ws.min_row

    min_column = ws.min_column

    max_row = ws.max_row

    max_column = ws.max_column

    for ic in range (min_column, max_column + 1):

        for ir in range(min_row, max_row + 1):

            current = ws.cell(ir, ic).value

            if current != None:

                l_all.append(str(current))

                l_limitc.append(ic)

                l_limitr.append(ir)

    l_limit.append(len(l_all))

stop = 0

for i in range (0, len(list_to_search)):

    for i2 in range (0, len(l_all)):

        if list_to_search[i] == l_all[i2]:

            for i3 in range(0, len(l_limit)):

                if i2 < l_limit[i3] and stop == 0:

                    nb_s = i3

                    stop = 1

            stop = 0

            print("Match found in sheet", nb_s + 1, "at column", l_limitc[i2], "at row", l_limitr[i2], "with", list_to_search[i])


